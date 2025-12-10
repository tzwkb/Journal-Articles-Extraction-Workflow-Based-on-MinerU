"""
主流程脚本 - 修复版
修复：
1. 图片路径处理
2. MinerU输出位置
3. 图片复制逻辑
"""

import yaml
import sys
from pathlib import Path
from concurrent.futures import ProcessPoolExecutor, as_completed
from jinja2 import Template
import shutil

from mineru_client import MinerUClient, FileTask, TaskState
from mineru_parser import MinerUParser
from article_translator import ArticleTranslator
from logger import Logger
from format_converter import FormatConverter
from outline_generator import OutlineGenerator
from path_manager import PathManager

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    from tqdm import tqdm
except ImportError:
    tqdm = None


class DocumentProcessor:
    """文档处理主类"""

    def __init__(self, config_path="config.yaml"):
        """
        初始化文档处理器

        Args:
            config_path: 配置文件路径
        """
        # 加载配置
        with open(config_path, 'r', encoding='utf-8') as f:
            self.config = yaml.safe_load(f)

        self.logger = Logger()
        self.output_base = Path(self.config['paths']['output_base'])

        # 初始化MinerU客户端
        self.mineru = MinerUClient(
            api_token=self.config['api']['mineru_token'],
            verify_ssl=False,
            max_retries=self.config['retry']['mineru_max_retries']
        )

        # 初始化解析器（修改输出目录到output/MinerU）
        mineru_output_dir = self.output_base / self.config['output']['mineru_folder']
        self.parser = MinerUParser(output_dir=str(mineru_output_dir))

        # 初始化格式转换器
        self.converter = FormatConverter(self.config, self.logger, self.output_base)

        # 初始化大纲生成器
        self.outline_gen = OutlineGenerator(self.config, self.logger, self.output_base)

        # 初始化路径管理器
        self.path_mgr = PathManager(self.config, self.logger)

        # 初始化文件夹结构
        self._init_directories()

    def _init_directories(self):
        """初始化所需的文件夹结构"""
        input_base = Path(self.config['paths']['input_base'])
        output_base = Path(self.config['paths']['output_base'])
        terminology_folder = Path(self.config['paths']['terminology_folder'])

        # 输出文件夹名称
        mineru_folder = self.config['output']['mineru_folder']
        html_folder = self.config['output']['html_folder']
        pdf_folder = self.config['output']['pdf_folder']
        docx_folder = self.config['output']['docx_folder']
        cache_folder = self.config['output']['cache_folder']

        # 创建所有必要的目录
        folders = [
            input_base,
            terminology_folder,
            output_base / mineru_folder,
            output_base / html_folder,
            output_base / pdf_folder,
            output_base / docx_folder,
            output_base / cache_folder / 'outlines',
        ]

        for folder in folders:
            folder.mkdir(parents=True, exist_ok=True)

        self.logger.info(f"文件夹结构初始化完成")

    def load_terminology_from_excel(self) -> dict:
        """
        从 terminology 文件夹下的 Excel 文件加载术语库

        Returns:
            术语字典 {"English": "中文"}
        """
        terminology_folder = Path(self.config['paths']['terminology_folder'])

        if not terminology_folder.exists():
            self.logger.warning(f"术语库文件夹不存在: {terminology_folder}")
            return {}

        if not load_workbook:
            self.logger.warning("openpyxl 未安装，无法读取 Excel 术语库")
            return {}

        glossary = {}
        excel_files = list(terminology_folder.glob("*.xlsx")) + list(terminology_folder.glob("*.xls"))

        if not excel_files:
            self.logger.warning(f"术语库文件夹中没有 Excel 文件: {terminology_folder}")
            return {}

        self.logger.info(f"正在加载术语库，共 {len(excel_files)} 个 Excel 文件...")

        for excel_file in excel_files:
            try:
                workbook = load_workbook(excel_file, read_only=True, data_only=True)

                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]

                    if sheet.max_row <= 1:
                        continue

                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if len(row) >= 2 and row[0] and row[1]:
                            english_term = str(row[0]).strip()
                            chinese_term = str(row[1]).strip()

                            if english_term and chinese_term:
                                glossary[english_term] = chinese_term

                workbook.close()
                self.logger.info(f"  已加载: {excel_file.name} - {len(glossary)} 个术语")

            except Exception as e:
                self.logger.error(f"加载 Excel 文件失败: {excel_file.name} - {str(e)}")

        self.logger.success(f"术语库加载完成，共 {len(glossary)} 个术语")
        return glossary

    def batch_process(self):
        """
        批量处理 input 文件夹中的所有 PDF 文件（多文件并发）
        """
        self.logger.info("=" * 60)
        self.logger.info("批量处理模式")
        self.logger.info("=" * 60)

        # 1. 扫描输入文件
        file_list = self.path_mgr.scan_input_files()

        if not file_list:
            self.logger.error("没有找到要处理的 PDF 文件")
            return

        # 2. 加载全局术语库（从 Excel）
        excel_glossary = self.load_terminology_from_excel()

        # 3. 多文件并发处理
        max_workers = self.config['concurrency']['max_files']
        self.logger.info(f"开始并发处理，并发数: {max_workers}")

        success_count = 0
        failure_count = 0
        results = []

        with ProcessPoolExecutor(max_workers=max_workers) as executor:
            future_to_file = {
                executor.submit(self._process_single_file, relative_path, pdf_path, excel_glossary):
                (relative_path, pdf_path)
                for relative_path, pdf_path in file_list
            }

            if tqdm:
                future_iterator = tqdm(as_completed(future_to_file), total=len(file_list), desc="处理进度")
            else:
                future_iterator = as_completed(future_to_file)

            for future in future_iterator:
                relative_path, pdf_path = future_to_file[future]
                try:
                    result = future.result()
                    if result['success']:
                        success_count += 1
                        self.logger.success(f"✓ 完成: {relative_path}")
                    else:
                        failure_count += 1
                        self.logger.error(f"✗ 失败: {relative_path} - {result.get('error', 'Unknown error')}")
                    results.append(result)
                except Exception as e:
                    failure_count += 1
                    self.logger.error(f"✗ 失败: {relative_path} - {str(e)}")
                    results.append({'success': False, 'file': relative_path, 'error': str(e)})

        # 4. 输出汇总
        self.logger.info("=" * 60)
        self.logger.info(f"批量处理完成！")
        self.logger.info(f"  成功: {success_count} 个文件")
        self.logger.info(f"  失败: {failure_count} 个文件")
        self.logger.info("=" * 60)

        return results

    def _process_single_file(self, relative_path: str, pdf_path: str, excel_glossary: dict) -> dict:
        """
        处理单个 PDF 文件（用于多进程调用）

        Args:
            relative_path: 相对路径
            pdf_path: PDF 绝对路径
            excel_glossary: Excel 术语库

        Returns:
            处理结果字典
        """
        try:
            output_paths = self.path_mgr.get_output_paths(relative_path)
            self.run(pdf_path, output_paths, excel_glossary)

            return {
                'success': True,
                'file': relative_path,
                'output_paths': {k: str(v) for k, v in output_paths.items()}
            }
        except Exception as e:
            return {
                'success': False,
                'file': relative_path,
                'error': str(e)
            }

    def run(self, pdf_path: str, output_paths: dict = None, excel_glossary: dict = None):
        """
        运行完整流程

        Args:
            pdf_path: PDF文件路径
            output_paths: 自定义输出路径字典（可选）
            excel_glossary: Excel术语库（可选）
        """
        self.logger.info("=" * 60)
        self.logger.info("开始处理文档")
        self.logger.info("=" * 60)

        try:
            # 步骤1: 生成大纲
            outline = self.outline_gen.generate_outline(pdf_path, output_paths)

            # 步骤2: MinerU解析
            content_list, extract_dir = self.parse_with_mineru(pdf_path, output_paths)

            # 步骤3: 使用 Excel 术语库（不使用 AI 生成的术语）
            combined_glossary = excel_glossary or {}
            
            if combined_glossary:
                self.logger.info(f"术语库加载完成: {len(combined_glossary)} 个术语")
            else:
                self.logger.warning("未找到术语库，将不进行术语预替换")

            # 步骤4: 初始化翻译器
            translator = ArticleTranslator(
                api_key=self.config['api']['translation_api_key'],
                api_url=self.config['api']['translation_api_base_url'],
                model=self.config['api']['translation_api_model'],
                glossary=combined_glossary,
                case_sensitive=False,
                whole_word_only=True,
                config=self.config
            )

            # 步骤5: 处理内容并翻译
            original_html, translated_html = self.process_content(
                content_list, outline, translator, extract_dir, output_paths
            )

            # 步骤6: 导出格式
            self.converter.export_formats(original_html, translated_html, output_paths)

            self.logger.info("=" * 60)
            self.logger.success("处理完成！")
            self.logger.info("=" * 60)

        except Exception as e:
            self.logger.error(f"处理失败: {str(e)}")
            import traceback
            traceback.print_exc()
            raise

    def parse_with_mineru(self, pdf_path: str, output_paths: dict = None) -> tuple:
        """
        使用MinerU解析PDF

        Args:
            pdf_path: PDF文件路径
            output_paths: 自定义输出路径字典（可选）

        Returns:
            (content_list, extract_dir) - 内容列表和解压目录
        """
        self.logger.info("\n>>> 步骤2: 使用MinerU解析PDF...")

        # 确定ZIP保存路径（output/MinerU/相对路径）
        if output_paths and 'mineru' in output_paths:
            expected_zip = Path(output_paths['mineru'])
        else:
            mineru_folder = self.config['output']['mineru_folder']
            mineru_dir = self.output_base / mineru_folder
            pdf_name = Path(pdf_path).stem
            expected_zip = mineru_dir / f"{pdf_name}_result.zip"

        expected_zip.parent.mkdir(parents=True, exist_ok=True)

        # 检查是否已有解析结果
        if expected_zip.exists():
            self.logger.info("发现已有MinerU解析结果，直接加载...")
            parsed = self.parser.parse_zip_result(
                str(expected_zip),
                source_file_name=Path(pdf_path).name
            )
            # 获取解压目录
            extract_dir = self.parser.output_dir / Path(expected_zip).stem
            self.logger.success(f"解析结果已加载: {len(parsed.json_content)} 个内容块")
            return parsed.json_content, str(extract_dir)

        # 上传并解析
        file_task = FileTask(
            file_name=Path(pdf_path).name,
            file_path=pdf_path,
            data_id=Path(pdf_path).stem
        )

        self.logger.info("正在上传PDF到MinerU...")
        batch_id, _ = self.mineru.batch_upload_files([file_task])

        self.logger.info("等待MinerU解析完成...")
        results = self.mineru.wait_for_completion(batch_id, poll_interval=10)

        # 下载结果到指定位置
        downloaded = self.mineru.download_all_results(results, str(expected_zip.parent))

        # 检查是否成功下载
        if not downloaded:
            error_msg = "MinerU解析失败，没有可下载的结果。"
            # 检查results中的失败原因
            for result in results:
                if result.state == TaskState.FAILED:
                    reason = result.err_msg or '未知原因'
                    error_msg += f"\n失败原因: {reason}"
                    error_msg += "\n\n可能的解决方案:"
                    error_msg += "\n1. 检查PDF文件是否损坏或加密"
                    error_msg += "\n2. 尝试重新下载或转换PDF文件"
                    error_msg += "\n3. 检查PDF文件大小是否超过限制"
            self.logger.error(error_msg)
            raise RuntimeError(error_msg)

        # 获取下载的zip文件路径
        zip_path = list(downloaded.values())[0]

        # 如果下载位置不是目标位置，移动文件
        if Path(zip_path) != expected_zip:
            shutil.move(zip_path, str(expected_zip))

        # 解析ZIP
        parsed = self.parser.parse_zip_result(
            str(expected_zip),
            source_file_name=Path(pdf_path).name
        )

        # 获取解压目录
        extract_dir = self.parser.output_dir / Path(expected_zip).stem

        self.logger.success(f"解析完成: {len(parsed.json_content)} 个内容块")
        return parsed.json_content, str(extract_dir)

    def process_content(
        self,
        content_list: list,
        outline: dict,
        translator: ArticleTranslator,
        extract_dir: str,
        output_paths: dict = None
    ) -> tuple:
        """
        处理内容并翻译

        Args:
            content_list: MinerU返回的content_list
            outline: 文档大纲
            translator: 翻译器实例
            extract_dir: MinerU解压目录
            output_paths: 输出路径字典

        Returns:
            (original_html, translated_html) 元组
        """
        self.logger.info("\n>>> 步骤3: 处理内容并翻译...")

        # 按页分组
        pages = {}
        for item in content_list:
            page_idx = item.get('page_idx', 0)
            if page_idx not in pages:
                pages[page_idx] = []
            pages[page_idx].append(item)

        self.logger.info(f"共 {len(pages)} 页")

        # 处理图片：复制到HTML目录并更新路径
        self._process_images(content_list, extract_dir, output_paths)

        # 收集翻译任务
        tasks = []
        for page_idx in sorted(pages.keys()):
            items = pages[page_idx]

            # 极简合并：处理连字符断词和跨列分割
            merged_items = self._merge_split_texts(items)

            # 获取章节上下文
            chapter_context = self._get_chapter_context(page_idx, outline)

            for idx, item in enumerate(merged_items):
                if item['type'] in ['header', 'footer', 'page_number']:
                    continue

                # 添加上下文窗口（前后100字符）
                context = chapter_context.copy()
                if idx > 0 and merged_items[idx - 1].get('text'):
                    context['prev_text'] = merged_items[idx - 1]['text'][-100:]
                else:
                    context['prev_text'] = ''

                if idx < len(merged_items) - 1 and merged_items[idx + 1].get('text'):
                    context['next_text'] = merged_items[idx + 1]['text'][:100]
                else:
                    context['next_text'] = ''

                if item['type'] == 'text' and item.get('text'):
                    tasks.append((item, 'text_zh', item['text'], context))

                if item['type'] == 'image' and item.get('image_caption'):
                    caption_text = ' '.join(item['image_caption'])
                    tasks.append((item, 'caption_zh', caption_text, context))

        self.logger.info(f"共收集 {len(tasks)} 个翻译任务，开始并发翻译...")

        # 批量并发翻译
        translation_tasks = [(text, context) for _, _, text, context in tasks]
        translations = translator.translate_batch(translation_tasks)

        # 赋值翻译结果
        for i, (item, field_name, _, _) in enumerate(tasks):
            translated_text = translations[i]

            # 检查是否是合并项
            if item.get('merged') and 'original_items' in item:
                # 拆分译文回原始TEXT块
                originals = item['original_items']

                # 按原始文本长度比例拆分
                len1 = len(originals[0]['text'])
                len2 = len(originals[1]['text'])
                total_len = len1 + len2

                if total_len > 0:
                    ratio = len1 / total_len
                    split_point = int(len(translated_text) * ratio)

                    # 分配译文
                    originals[0][field_name] = translated_text[:split_point].strip()
                    originals[1][field_name] = translated_text[split_point:].strip()

                    # 保留合并信息（用于调试）
                    originals[0]['_merged_from'] = item['text']
                    originals[1]['_merged_from'] = item['text']
                else:
                    # 异常情况：原始文本长度为0，直接赋值给第一个
                    originals[0][field_name] = translated_text
            else:
                # 未合并的项，直接赋值
                item[field_name] = translated_text

            if (i + 1) % max(1, len(tasks) // 10) == 0:
                progress = (i + 1) * 100 // len(tasks)
                self.logger.info(f"  翻译进度: {i + 1}/{len(tasks)} ({progress}%)")

        self.logger.success(f"翻译完成: {len(tasks)} 个内容块")

        # 生成HTML
        self.logger.info("正在生成HTML...")
        original_html = self._render_html(pages, language='en')
        translated_html = self._render_html(pages, language='zh')

        self.logger.success("HTML已生成")

        return original_html, translated_html

    def _process_images(self, content_list: list, extract_dir: str, output_paths: dict = None):
        """
        处理图片：复制图片到HTML输出目录并更新路径

        Args:
            content_list: 内容列表
            extract_dir: MinerU解压目录
            output_paths: 输出路径字典
        """
        extract_dir = Path(extract_dir)
        source_images_dir = extract_dir / "images"

        if not source_images_dir.exists():
            self.logger.warning(f"未找到图片目录: {source_images_dir}")
            return

        # 确定目标图片目录（统一放在 output/HTML/images/）
        html_folder = self.config['output']['html_folder']
        html_base_dir = self.output_base / html_folder
        
        if output_paths and 'html_original' in output_paths:
            # 使用与 HTML 文件相同的目录层级
            html_dir = Path(output_paths['html_original']).parent
        else:
            html_dir = html_base_dir

        target_images_dir = html_dir / "images"
        target_images_dir.mkdir(parents=True, exist_ok=True)

        self.logger.info(f"正在复制图片: {source_images_dir} -> {target_images_dir}")

        # 复制图片并更新路径
        copied_count = 0
        for item in content_list:
            if item.get('type') == 'image' and item.get('img_path'):
                img_rel_path = item['img_path']
                source_img = extract_dir / img_rel_path

                if source_img.exists():
                    img_filename = Path(img_rel_path).name
                    target_img = target_images_dir / img_filename

                    # 复制图片
                    shutil.copy2(source_img, target_img)

                    # 更新路径：
                    # 1. 相对路径用于 HTML（images/xxx.jpg）
                    # 2. 绝对路径用于 PDF/DOCX 转换（存储在 img_path_absolute）
                    item['img_path'] = f"images/{img_filename}"
                    item['img_path_absolute'] = str(target_img.absolute())
                    copied_count += 1
                else:
                    self.logger.warning(f"图片文件不存在: {source_img}")

        if copied_count > 0:
            self.logger.success(f"已复制 {copied_count} 张图片")
        else:
            self.logger.warning("未找到任何图片文件")

    def _merge_split_texts(self, items: list) -> list:
        """
        极简合并 - 只处理明确的TEXT分割

        规则1: 连字符断词 (如 "frig-" + "ates")
        规则2: 跨列无标点 (如左列 "...limestone" + 右列 "V pedestal")
        规则3: 同列分割 (如 "...Pound" + "force was...")

        Args:
            items: 单页的内容项列表

        Returns:
            合并后的内容项列表（保留original_items字段）
        """
        merged = []
        i = 0

        while i < len(items):
            current = items[i]

            # 只处理text类型
            if current.get('type') != 'text' or not current.get('text'):
                merged.append(current)
                i += 1
                continue

            # 检查是否与下一项合并
            should_merge = False
            if i + 1 < len(items):
                next_item = items[i + 1]

                # 下一项也必须是text
                if next_item.get('type') == 'text' and next_item.get('text'):
                    # 同一页
                    if current.get('page_idx') == next_item.get('page_idx'):
                        text1 = current['text'].strip()
                        bbox1 = current.get('bbox', [0, 0, 0, 0])
                        bbox2 = next_item.get('bbox', [0, 0, 0, 0])

                        # 规则1: 连字符结尾 (100%确定是断词)
                        if text1.endswith('-'):
                            should_merge = True
                        # 规则2: 跨列 + 无句末标点
                        elif bbox2[0] - bbox1[2] > 80:  # x间距 > 80像素（跨列）
                            if text1 and text1[-1] not in '.!?。！？':
                                should_merge = True
                        # 规则3: 同列内分割 - text1无标点结尾 + text2小写开头
                        else:
                            text2 = next_item['text'].strip()
                            # text1不以标点结尾 且 text2以小写字母开头
                            if (text1 and text1[-1] not in '.!?。！？,;:' and
                                text2 and text2[0].islower()):
                                should_merge = True

            if should_merge:
                # 合并两个TEXT块
                merged_item = current.copy()
                merged_item['text'] = current['text'].rstrip() + ' ' + next_item['text'].lstrip()
                merged_item['original_items'] = [current, next_item]
                merged_item['merged'] = True
                merged.append(merged_item)
                i += 2  # 跳过下一项
            else:
                merged.append(current)
                i += 1

        return merged

    def _get_chapter_context(self, page_idx: int, outline: dict) -> dict:
        """获取页面对应的章节上下文"""
        # 确保 page_idx 是整数
        try:
            page_num = int(page_idx)
        except (ValueError, TypeError):
            return {}

        for chapter in outline.get('structure', []):
            pages = chapter.get('pages', [])
            if len(pages) >= 2:
                try:
                    # 确保 start 和 end 也是整数
                    start = int(pages[0])
                    end = int(pages[1])
                    if start <= page_num <= end:
                        return {
                            'chapter_title': chapter.get('title', ''),
                            'chapter_summary': chapter.get('summary', ''),
                            'keywords': chapter.get('keywords', [])
                        }
                except (ValueError, TypeError, IndexError):
                    continue
        return {}

    def _render_html(self, pages: dict, language: str) -> str:
        """渲染HTML"""
        with open('page_template.html', 'r', encoding='utf-8') as f:
            template = Template(f.read())

        return template.render(pages=pages, language=language)


def main():
    """命令行入口"""
    if len(sys.argv) == 1:
        interactive_mode()
        return

    if sys.argv[1] in ["--batch", "-b", "--interactive", "-i"]:
        interactive_mode()
    else:
        print(f"❌ 未知参数: {sys.argv[1]}")
        print("使用 'python main.py -h' 查看帮助")
        sys.exit(1)

def interactive_mode():
    """交互式命令行界面"""
    processor = DocumentProcessor()

    while True:
        print("\n" + "="*60)
        print("  MinerU 文档翻译工具 - 交互模式")
        print("="*60)
        print("\n请选择操作：")
        print("  [1] 批量处理（递归扫描 input/ 文件夹）")
        print("  [2] 查看配置信息")
        print("  [3] 查看输入文件列表")
        print("  [4] 清除缓存")
        print("  [0] 退出")
        print()

        choice = input("请输入选项 [0-4]: ").strip()

        if choice == "0":
            print("\n再见！")
            break
        elif choice == "1":
            batch_mode_interactive(processor)
        elif choice == "2":
            show_config(processor)
        elif choice == "3":
            show_input_files(processor)
        elif choice == "4":
            clear_cache(processor)
        else:
            print("❌ 无效选项，请重新选择")


def batch_mode_interactive(processor):
    """批量处理交互模式"""
    print("\n" + "-"*60)
    print("  批量处理模式")
    print("-"*60)

    file_list = processor.path_mgr.scan_input_files()

    if not file_list:
        print("\n❌ input/ 文件夹中没有找到 PDF 文件")
        print("   请先将 PDF 文件放入 input/ 文件夹")
        input("\n按回车键继续...")
        return

    print(f"\n找到 {len(file_list)} 个 PDF 文件:")
    for i, (rel_path, abs_path) in enumerate(file_list[:10], 1):
        print(f"  {i}. {rel_path}")

    if len(file_list) > 10:
        print(f"  ... 还有 {len(file_list) - 10} 个文件")

    print(f"\n并发配置:")
    print(f"  - 文件并发数: {processor.config['concurrency']['max_files']}")
    print(f"  - 翻译并发数: {processor.config['concurrency']['initial_translation_workers']} (初始)")

    confirm = input(f"\n确认开始批量处理？[y/N]: ").strip().lower()

    if confirm != 'y':
        print("已取消")
        return

    try:
        print("\n开始批量处理...")
        processor.batch_process()
        print("\n✓ 批量处理完成！")
    except Exception as e:
        print(f"\n❌ 批量处理失败: {str(e)}")

    input("\n按回车键继续...")


def show_config(processor):
    """显示配置信息"""
    print("\n" + "-"*60)
    print("  当前配置信息")
    print("-"*60)

    config = processor.config

    print("\n📡 API 配置:")
    print(f"  MinerU Token: {'已配置' if config['api']['mineru_token'] != 'YOUR_MINERU_TOKEN' else '❌ 未配置'}")
    print(f"  Translation API: {config['api']['translation_api_base_url']}")
    print(f"  Translation Model: {config['api']['translation_api_model']}")

    print("\n🔄 并发配置:")
    print(f"  文件并发数: {config['concurrency']['max_files']}")
    print(f"  翻译并发数: {config['concurrency']['initial_translation_workers']}-{config['concurrency']['max_translation_workers']}")

    input("\n按回车键继续...")


def show_input_files(processor):
    """显示输入文件列表"""
    print("\n" + "-"*60)
    print("  输入文件列表")
    print("-"*60)

    file_list = processor.path_mgr.scan_input_files()

    if not file_list:
        print("\n❌ input/ 文件夹中没有找到 PDF 文件")
    else:
        print(f"\n找到 {len(file_list)} 个 PDF 文件:\n")
        for i, (rel_path, abs_path) in enumerate(file_list, 1):
            file_size = Path(abs_path).stat().st_size / (1024 * 1024)
            print(f"  {i:3d}. {rel_path:50s} ({file_size:.1f} MB)")

    input("\n按回车键继续...")


def clear_cache(processor):
    """清除缓存"""
    print("\n" + "-"*60)
    print("  清除缓存")
    print("-"*60)

    cache_dir = processor.output_base / "cache"

    if not cache_dir.exists():
        print("\n没有缓存需要清除")
        input("\n按回车键继续...")
        return

    total_size = 0
    file_count = 0
    for file in cache_dir.rglob("*"):
        if file.is_file():
            total_size += file.stat().st_size
            file_count += 1

    print(f"\n缓存统计:")
    print(f"  文件数: {file_count}")
    print(f"  总大小: {total_size / (1024 * 1024):.1f} MB")

    confirm = input("\n确认清除所有缓存？[y/N]: ").strip().lower()

    if confirm != 'y':
        print("已取消")
        input("\n按回车键继续...")
        return

    try:
        shutil.rmtree(cache_dir)
        cache_dir.mkdir(parents=True, exist_ok=True)
        print("\n✓ 缓存已清除")
    except Exception as e:
        print(f"\n❌ 清除失败: {str(e)}")

    input("\n按回车键继续...")


if __name__ == "__main__":
    main()